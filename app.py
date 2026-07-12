# app.py
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

st.set_page_config(
    page_title="安全库存管理系统",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
    <style>
    .main-header {
        font-size: 2.2rem;
        font-weight: 700;
        color: #1f77b4;
        margin-bottom: 0.5rem;
        text-align: center;
    }
    .file-status {
        padding: 0.25rem 0.5rem;
        border-radius: 0.25rem;
        font-size: 0.9rem;
    }
    .file-uploaded {
        background-color: #d4edda;
        color: #155724;
    }
    .file-missing {
        background-color: #f8d7da;
        color: #721c24;
    }
    .file-optional {
        background-color: #fff3cd;
        color: #856404;
    }
    .monthly-update-badge {
        background-color: #28a745;
        color: white;
        padding: 0.2rem 0.8rem;
        border-radius: 1rem;
        font-size: 0.8rem;
        display: inline-block;
    }
    </style>
""", unsafe_allow_html=True)

@st.cache_data
def load_safety_stock(file):
    """加载安全库存模板"""
    try:
        df = pd.read_excel(file)
        if '物料编码' in df.columns:
            df['物料编码'] = df['物料编码'].astype(str)
        return df
    except Exception as e:
        st.error(f"加载安全库存文件失败: {e}")
        return None

@st.cache_data
def load_inventory(file):
    """加载6.1库存文件"""
    try:
        df = pd.read_excel(file, sheet_name='Data')
        df_qualified = df[df['存储地点'] == 3006].copy()
        inventory_summary = df_qualified.groupby('物料')['非限制使用的库存'].sum().reset_index()
        inventory_summary.columns = ['物料编码', '实际库存']
        inventory_summary['物料编码'] = inventory_summary['物料编码'].astype(str)
        return inventory_summary
    except Exception as e:
        st.error(f"加载库存文件失败: {e}")
        return None

@st.cache_data
def load_usage(file):
    """加载月度使用量文件"""
    try:
        df = pd.read_excel(file)
        if '物料编码' in df.columns and '日期' in df.columns:
            df['物料编码'] = df['物料编码'].astype(str)
            # 提取月份信息，计算日均消耗
            df['月份'] = pd.to_datetime(df['日期']).dt.month
            monthly_usage = df.groupby(['物料编码', '月份'])['消耗量'].sum().reset_index()
            # 按月份计算日均消耗（假设当月天数）
            days_in_month = {1:31, 2:28, 3:31, 4:30, 5:31, 6:30, 
                           7:31, 8:31, 9:30, 10:31, 11:30, 12:31}
            monthly_usage['日均消耗'] = monthly_usage.apply(
                lambda x: x['消耗量'] / days_in_month.get(x['月份'], 30), axis=1
            )
            # 取最近一个月的日均消耗
            latest_month = monthly_usage.groupby('物料编码')['月份'].max().reset_index()
            latest_month.columns = ['物料编码', '最新月份']
            result = latest_month.merge(monthly_usage, on=['物料编码', '月份'], how='left')
            return result[['物料编码', '日均消耗']]
        else:
            return None
    except Exception as e:
        st.error(f"加载使用量文件失败: {e}")
        return None

@st.cache_data
def load_forecast(file):
    """加载预测库存文件"""
    try:
        df = pd.read_excel(file)
        if '物料编码' in df.columns:
            df['物料编码'] = df['物料编码'].astype(str)
            forecast_col = [col for col in df.columns if '预测' in col or '库存' in col][0]
            forecast_summary = df[['物料编码', forecast_col]].copy()
            forecast_summary.columns = ['物料编码', '预测库存']
            return forecast_summary
        else:
            return None
    except Exception as e:
        st.error(f"加载预测文件失败: {e}")
        return None

def calculate_safety_stock(df):
    """计算安全库存 - 支持月度滚动更新"""
    df_calc = df.copy()
    
    if '物料编码' in df_calc.columns:
        df_calc['物料编码'] = df_calc['物料编码'].astype(str)
    
    month_cols = ['M01', 'M02', 'M03', 'M04', 'M05', 'M06', 'M07', 'M08', 'M09']
    for col in month_cols:
        if col in df_calc.columns:
            df_calc[col] = df_calc[col].abs()
    
    # 1. 未来3个月月均用量（M07-M09）
    df_calc['未来3个月月均用量'] = df_calc[['M07', 'M08', 'M09']].mean(axis=1)
    
    # 2. 过去6个月月均用量（M01-M06）
    df_calc['过去6个月月均用量'] = df_calc[['M01', 'M02', 'M03', 'M04', 'M05', 'M06']].mean(axis=1)
    
    # 3. 质量风险系数映射
    quality_risk_map = {'高': 1.5, '中': 1.2, '低': 1.0}
    if '质量风险系数' in df_calc.columns:
        df_calc['质量风险系数'] = df_calc['质量风险系数'].map(quality_risk_map).fillna(1.0)
    else:
        df_calc['质量风险系数'] = 1.0
    
    # 4. 品类策略系数映射
    category_map = {'A类': 1.3, 'B类': 1.1, 'C类': 0.9, 'D类': 0.7}
    if '品类策略系数' in df_calc.columns:
        df_calc['品类策略系数'] = df_calc['品类策略系数'].map(category_map).fillna(1.0)
    else:
        df_calc['品类策略系数'] = 1.0
    
    # 5. 采购周期系数查表
    lead_time_map = {'1周': 1.2, '2周': 1.4, '3周': 1.6, '4周': 1.8, '5周': 2.0, '6周': 2.2}
    if '平均交货周期' in df_calc.columns:
        df_calc['采购周期系数'] = df_calc['平均交货周期'].map(lead_time_map).fillna(1.5)
    else:
        df_calc['采购周期系数'] = 1.5
    
    # 6. 计算安全库存量
    df_calc['安全库存量'] = (
        df_calc['未来3个月月均用量'] * 
        df_calc['采购周期系数'] * 
        df_calc['质量风险系数'] * 
        df_calc['品类策略系数']
    ).round(2)
    
    return df_calc

def load_mock_data():
    """生成模拟数据（用于演示）"""
    np.random.seed(42)
    materials = [f'MAT-{i:04d}' for i in range(1, 51)]
    data = []
    for mat in materials:
        m01_m06 = -np.random.uniform(100, 500, 6)
        m07_m09 = -np.random.uniform(150, 600, 3)
        row = {
            '物料编码': mat,
            '物料描述': f'物料_{mat}',
            '工厂': np.random.choice(['广州工厂', '上海工厂', '北京工厂']),
            '存储地点': np.random.choice(['A区', 'B区', 'C区']),
            '单位': 'KG',
            'M01': m01_m06[0], 'M02': m01_m06[1], 'M03': m01_m06[2],
            'M04': m01_m06[3], 'M05': m01_m06[4], 'M06': m01_m06[5],
            'M07': m07_m09[0], 'M08': m07_m09[1], 'M09': m07_m09[2],
            '平均交货周期': np.random.choice(['1周', '2周', '3周', '4周']),
            '质量风险系数': np.random.choice(['高', '中', '低'], p=[0.2, 0.5, 0.3]),
            '品类策略系数': np.random.choice(['A类', 'B类', 'C类'], p=[0.2, 0.4, 0.4]),
        }
        data.append(row)
    return pd.DataFrame(data)

def export_to_excel(df, template_columns=None):
    """导出Excel，保持与模板完全一致的格式"""
    output = io.BytesIO()
    
    if template_columns and len(template_columns) > 0:
        export_cols = [col for col in template_columns if col in df.columns]
        if len(export_cols) == len(template_columns):
            export_df = df[export_cols].copy()
        else:
            export_df = df.copy()
    else:
        export_df = df.copy()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        export_df.to_excel(writer, sheet_name='计算结果', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['计算结果']
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        number_cols = ['安全库存量', '未来3个月月均用量', '过去6个月月均用量']
        for col in number_cols:
            if col in export_df.columns:
                col_idx = export_df.columns.get_loc(col) + 1
                for row in range(2, len(export_df) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.number_format = '#,##0.00'
    
    return output.getvalue()

def main():
    # 显示标题
    st.markdown('<div class="main-header">📦 安全库存管理系统</div>', unsafe_allow_html=True)
    
    # 显示月度更新标识
    current_month = datetime.now().strftime('%Y年%m月')
    st.markdown(f'<span class="monthly-update-badge">📅 {current_month} 月度报告</span>', unsafe_allow_html=True)
    st.markdown("---")
    
    # 初始化session state
    if 'safety_data' not in st.session_state:
        st.session_state.safety_data = None
    if 'template_columns' not in st.session_state:
        st.session_state.template_columns = None
    if 'inventory_data' not in st.session_state:
        st.session_state.inventory_data = None
    if 'usage_data' not in st.session_state:
        st.session_state.usage_data = None
    if 'forecast_data' not in st.session_state:
        st.session_state.forecast_data = None
    if 'last_update' not in st.session_state:
        st.session_state.last_update = None
    
    # 侧边栏
    with st.sidebar:
        st.markdown("## 📂 数据上传")
        st.caption("上传您的Excel文件，系统将自动计算安全库存")
        
        # 显示上次更新时间
        if st.session_state.last_update:
            st.info(f"🕐 上次更新: {st.session_state.last_update}")
        
        # 示例数据按钮
        if st.button("📊 加载示例数据"):
            st.session_state.safety_data = calculate_safety_stock(load_mock_data())
            st.session_state.template_columns = load_mock_data().columns.tolist()
            st.session_state.last_update = datetime.now().strftime('%Y-%m-%d %H:%M')
            st.rerun()
        
        st.markdown("---")
        
        st.markdown("### 🔴 必须上传")
        safety_stock_file = st.file_uploader(
            "① 安全库存（需计算）.xlsx",
            type=['xlsx', 'xls'],
            key='safety_stock_upload'
        )
        if safety_stock_file:
            st.markdown(f'<span class="file-status file-uploaded">✅ 已上传: {safety_stock_file.name}</span>', unsafe_allow_html=True)
        else:
            st.markdown('<span class="file-status file-missing">❌ 未上传</span>', unsafe_allow_html=True)
        
        st.markdown("---")
        
        st.markdown("### 🟡 推荐上传（月度更新）")
        st.caption("💡 每月1号更新此文件")
        inventory_file = st.file_uploader(
            "② X.1库存-去除名称.xlsx",
            type=['xlsx', 'xls'],
            key='inventory_upload'
        )
        if inventory_file:
            st.markdown(f'<span class="file-status file-uploaded">✅ 已上传: {inventory_file.name}</span>', unsafe_allow_html=True)
        else:
            st.markdown('<span class="file-status file-optional">⏭️ 将生成模拟数据</span>', unsafe_allow_html=True)
        
        st.markdown("---")
        
        st.markdown("### 🟢 可选上传（月度更新）")
        st.caption("💡 每月初更新上月消耗数据")
        usage_file = st.file_uploader(
            "③ X月使用量.xlsx",
            type=['xlsx', 'xls'],
            key='usage_upload'
        )
        if usage_file:
            st.markdown(f'<span class="file-status file-uploaded">✅ 已上传: {usage_file.name}</span>', unsafe_allow_html=True)
        else:
            st.markdown('<span class="file-status file-optional">⏭️ 未上传</span>', unsafe_allow_html=True)
        
        st.markdown("---")
        
        st.markdown("### 🟢 可选上传（月度更新）")
        st.caption("💡 每月初更新预测数据")
        forecast_file = st.file_uploader(
            "④ 预测-X.1.xlsx",
            type=['xlsx', 'xls'],
            key='forecast_upload'
        )
        if forecast_file:
            st.markdown(f'<span class="file-status file-uploaded">✅ 已上传: {forecast_file.name}</span>', unsafe_allow_html=True)
        else:
            st.markdown('<span class="file-status file-optional">⏭️ 未上传</span>', unsafe_allow_html=True)
        
        st.markdown("---")
        
        # 导出按钮
        if st.session_state.safety_data is not None:
            st.download_button(
                label="📥 导出安全库存计算结果",
                data=export_to_excel(
                    st.session_state.safety_data,
                    st.session_state.template_columns
                ),
                file_name=f"安全库存计算结果_{datetime.now().strftime('%Y%m')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        st.markdown("---")
        st.caption("📅 月度更新流程")
        st.caption("1. 更新 M01~M09 数据")
        st.caption("2. 上传新库存快照")
        st.caption("3. 上传上月消耗明细")
        st.caption("4. 导出月度报告")
    
    # ========================================================================
    # 数据处理
    # ========================================================================
    
    # 加载安全库存数据
    if safety_stock_file is not None:
        try:
            df = pd.read_excel(safety_stock_file)
            if '物料编码' in df.columns:
                df['物料编码'] = df['物料编码'].astype(str)
            st.session_state.template_columns = df.columns.tolist()
            st.session_state.safety_data = calculate_safety_stock(df)
            st.session_state.last_update = datetime.now().strftime('%Y-%m-%d %H:%M')
            st.sidebar.success(f"✅ 已加载 {len(st.session_state.safety_data)} 条物料数据")
        except Exception as e:
            st.error(f"数据加载失败: {e}")
            df = load_mock_data()
            st.session_state.template_columns = df.columns.tolist()
            st.session_state.safety_data = calculate_safety_stock(df)
    else:
        if st.session_state.safety_data is None:
            df = load_mock_data()
            st.session_state.template_columns = df.columns.tolist()
            st.session_state.safety_data = calculate_safety_stock(df)
            st.sidebar.info("ℹ️ 使用模拟数据，点击「加载示例数据」刷新")
    
    # 加载其他数据
    if inventory_file is not None:
        st.session_state.inventory_data = load_inventory(inventory_file)
    if usage_file is not None:
        st.session_state.usage_data = load_usage(usage_file)
    if forecast_file is not None:
        st.session_state.forecast_data = load_forecast(forecast_file)
    
    # ========================================================================
    # 显示结果
    # ========================================================================
    
    if st.session_state.safety_data is not None:
        df_display = st.session_state.safety_data.copy()
        df_display['物料编码'] = df_display['物料编码'].astype(str)
        
        # 合并库存数据
        if st.session_state.inventory_data is not None:
            st.session_state.inventory_data['物料编码'] = st.session_state.inventory_data['物料编码'].astype(str)
            df_display = df_display.merge(st.session_state.inventory_data, on='物料编码', how='left')
            df_display['实际库存'] = df_display['实际库存'].fillna(0)
        else:
            np.random.seed(42)
            df_display['实际库存'] = df_display['安全库存量'] * np.random.uniform(0.5, 1.8, len(df_display))
            df_display['实际库存'] = df_display['实际库存'].round(2)
        
        # 合并预测数据
        if st.session_state.forecast_data is not None:
            st.session_state.forecast_data['物料编码'] = st.session_state.forecast_data['物料编码'].astype(str)
            df_display = df_display.merge(st.session_state.forecast_data, on='物料编码', how='left')
            df_display['预测库存'] = df_display['预测库存'].fillna(0)
        else:
            np.random.seed(42)
            df_display['预测库存'] = df_display['安全库存量'] * np.random.uniform(0.6, 1.5, len(df_display))
            df_display['预测库存'] = df_display['预测库存'].round(2)
        
        # 计算库存差异
        df_display['库存差异(实际-安全)'] = df_display['实际库存'] - df_display['安全库存量']
        df_display['库存状态'] = df_display['库存差异(实际-安全)'].apply(
            lambda x: '✅ 充足' if x >= 0 else '⚠️ 预警'
        )
        
        # 合并使用量数据
        if st.session_state.usage_data is not None:
            st.session_state.usage_data['物料编码'] = st.session_state.usage_data['物料编码'].astype(str)
            df_display = df_display.merge(st.session_state.usage_data, on='物料编码', how='left')
            df_display['日均消耗'] = df_display['日均消耗'].fillna(df_display['过去6个月月均用量'] / 30)
        else:
            df_display['日均消耗'] = df_display['过去6个月月均用量'] / 30
        
        # ====================================================================
        # Tab布局
        # ====================================================================
        tab1, tab2, tab3, tab4 = st.tabs([
            "📋 数据概览",
            "📊 库存状态分析",
            "⚠️ 预警列表",
            "📈 消耗趋势"
        ])
        
        # --- Tab 1: 数据概览 ---
        with tab1:
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("📦 物料总数", len(df_display))
            with col2:
                st.metric("📊 总安全库存", f"{df_display['安全库存量'].sum():,.0f}")
            with col3:
                st.metric("📈 平均月用量", f"{df_display['未来3个月月均用量'].mean():,.0f}")
            with col4:
                warning_count = len(df_display[df_display['库存状态'] == '⚠️ 预警'])
                st.metric("⚠️ 预警物料", warning_count, delta_color="inverse")
            
            st.markdown("---")
            
            # 筛选器
            col1, col2, col3 = st.columns(3)
            with col1:
                material_filter = st.text_input("🔍 筛选物料编码", placeholder="输入物料编码")
            with col2:
                if '工厂' in df_display.columns:
                    plant_options = ['全部'] + df_display['工厂'].unique().tolist()
                    plant_filter = st.selectbox("🏭 筛选工厂", plant_options)
                else:
                    plant_filter = '全部'
            with col3:
                if '库存状态' in df_display.columns:
                    status_options = ['全部'] + df_display['库存状态'].unique().tolist()
                    status_filter = st.selectbox("📊 筛选库存状态", status_options)
                else:
                    status_filter = '全部'
            
            # 应用筛选
            filtered_data = df_display.copy()
            if material_filter:
                filtered_data = filtered_data[filtered_data['物料编码'].str.contains(material_filter, case=False)]
            if plant_filter != '全部' and '工厂' in filtered_data.columns:
                filtered_data = filtered_data[filtered_data['工厂'] == plant_filter]
            if status_filter != '全部' and '库存状态' in filtered_data.columns:
                filtered_data = filtered_data[filtered_data['库存状态'] == status_filter]
            
            # 显示数据表
            display_cols = ['物料编码', '物料描述', '工厂', '存储地点', '单位',
                           '未来3个月月均用量', '过去6个月月均用量', '日均消耗',
                           '安全库存量', '实际库存', '预测库存',
                           '库存差异(实际-安全)', '库存状态']
            display_cols = [col for col in display_cols if col in filtered_data.columns]
            
            st.dataframe(
                filtered_data[display_cols],
                use_container_width=True,
                height=400
            )
        
        # --- Tab 2: 库存状态分析 ---
        with tab2:
            st.markdown("### 📊 库存对比分析")
            
            top_n = st.slider("显示物料数量", 10, 50, 20)
            plot_df = df_display.head(top_n).copy()
            
            fig = go.Figure()
            fig.add_trace(go.Bar(
                x=plot_df['物料编码'],
                y=plot_df['安全库存量'],
                name='安全库存',
                marker_color='#1f77b4'
            ))
            fig.add_trace(go.Bar(
                x=plot_df['物料编码'],
                y=plot_df['实际库存'],
                name='实际库存',
                marker_color='#ff7f0e'
            ))
            fig.add_trace(go.Bar(
                x=plot_df['物料编码'],
                y=plot_df['预测库存'],
                name='预测库存',
                marker_color='#2ca02c'
            ))
            
            fig.update_layout(
                title="安全库存 vs 实际库存 vs 预测库存",
                xaxis_title="物料编码",
                yaxis_title="库存量",
                barmode='group',
                height=400,
                showlegend=True
            )
            st.plotly_chart(fig, use_container_width=True)
            
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("#### 库存状态分布")
                status_counts = df_display['库存状态'].value_counts()
                fig_pie = px.pie(
                    values=status_counts.values,
                    names=status_counts.index,
                    title="库存状态分布",
                    color=status_counts.index,
                    color_discrete_map={'✅ 充足': '#2ca02c', '⚠️ 预警': '#d62728'}
                )
                st.plotly_chart(fig_pie, use_container_width=True)
            
            with col2:
                st.markdown("#### 日均消耗TOP10")
                top_usage = df_display.nlargest(10, '日均消耗')
                fig_bar = px.bar(
                    top_usage,
                    x='物料编码',
                    y='日均消耗',
                    title="日均消耗TOP10",
                    color='日均消耗',
                    color_continuous_scale='Viridis'
                )
                st.plotly_chart(fig_bar, use_container_width=True)
        
        # --- Tab 3: 预警列表 ---
        with tab3:
            st.markdown("### ⚠️ 预警列表")
            
            warning_df = df_display[df_display['库存状态'] == '⚠️ 预警'].copy()
            
            if not warning_df.empty:
                warning_df = warning_df.sort_values('库存差异(实际-安全)')
                st.markdown(f"**共 {len(warning_df)} 个物料低于安全库存**")
                
                display_cols = ['物料编码', '物料描述', '工厂', '存储地点',
                               '安全库存量', '实际库存', '库存差异(实际-安全)']
                display_cols = [col for col in display_cols if col in warning_df.columns]
                
                st.dataframe(
                    warning_df[display_cols],
                    use_container_width=True,
                    height=400
                )
                
                if st.button("📥 导出预警列表"):
                    csv = warning_df[display_cols].to_csv(index=False).encode('utf-8-sig')
                    st.download_button(
                        label="下载CSV",
                        data=csv,
                        file_name=f"安全库存预警_{datetime.now().strftime('%Y%m')}.csv",
                        mime="text/csv"
                    )
            else:
                st.success("🎉 所有物料库存充足，没有预警！")
        
        # --- Tab 4: 消耗趋势 ---
        with tab4:
            st.markdown("### 📈 消耗趋势")
            
            view_type = st.radio(
                "查看方式",
                ['所有物料总趋势', '单个物料趋势'],
                horizontal=True
            )
            
            if view_type == '所有物料总趋势':
                month_cols = ['M01', 'M02', 'M03', 'M04', 'M05', 'M06', 'M07', 'M08', 'M09']
                month_cols = [col for col in month_cols if col in df_display.columns]
                monthly_total = df_display[month_cols].abs().sum()
                
                fig = px.line(
                    x=month_cols,
                    y=monthly_total.values,
                    title="所有物料总消耗趋势",
                    labels={'x': '月份', 'y': '总消耗量'},
                    markers=True
                )
                fig.update_layout(height=400)
                st.plotly_chart(fig, use_container_width=True)
            else:
                material_list = df_display['物料编码'].unique().tolist()
                selected_material = st.selectbox("选择物料", material_list)
                
                if selected_material:
                    material_data = df_display[df_display['物料编码'] == selected_material]
                    month_cols = ['M01', 'M02', 'M03', 'M04', 'M05', 'M06', 'M07', 'M08', 'M09']
                    month_cols = [col for col in month_cols if col in material_data.columns]
                    values = material_data[month_cols].abs().values.flatten()
                    
                    fig = px.line(
                        x=month_cols,
                        y=values,
                        title=f"物料 {selected_material} 消耗趋势",
                        labels={'x': '月份', 'y': '消耗量'},
                        markers=True
                    )
                    fig.update_layout(height=400)
                    st.plotly_chart(fig, use_container_width=True)
                    
                    # 显示物料详情
                    detail_cols = ['物料编码', '物料描述', '工厂', '存储地点', 
                                  '未来3个月月均用量', '过去6个月月均用量', '安全库存量']
                    detail_cols = [col for col in detail_cols if col in material_data.columns]
                    st.dataframe(material_data[detail_cols])

if __name__ == "__main__":
    main()
